import csv
import logging
import os
import tempfile
import uuid
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

HEADER_MAP = {
    "word_phrase": ["word", "phrase", "word/phrase", "term", "vocabulary"],
    "definition": ["definition", "def", "meaning"],
    "synonyms": ["synonym", "syn"],
    "collocations": ["collocation", "coll", "common collocations"],
    "example": ["example", "sample sentence", "sentence"],
    "cefr_level": ["cefr", "level", "cefr level", "cefr_level"],
}


def _match_header(header: str) -> str | None:
    h = header.lower().strip()
    for field, patterns in HEADER_MAP.items():
        for pattern in patterns:
            if pattern in h:
                return field
    return None


def _detect_csv_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        return ","


def _process_rows(raw_headers: list[str], data_rows: list) -> dict:
    column_map: dict[int, str] = {}
    for i, header in enumerate(raw_headers):
        if not header:
            continue
        field = _match_header(header)
        if field and field not in column_map.values():
            column_map[i] = field

    if "word_phrase" not in column_map.values():
        found = ", ".join(f'"{h}"' for h in raw_headers if h)
        return {
            "error": (
                "❌ Could not find a Word/Phrase column. "
                f"Found headers: {found}\n\n"
                "Make sure your first column header contains 'word' or 'phrase'."
            )
        }

    entries: list[dict] = []
    for row in data_rows:
        cells = list(row)
        entry: dict[str, str | None] = {}
        for col_idx, field_name in column_map.items():
            if col_idx < len(cells) and cells[col_idx] is not None:
                val = str(cells[col_idx]).strip()
                entry[field_name] = val if val else None
            else:
                entry[field_name] = None
        if entry.get("word_phrase"):
            entries.append(entry)

    if not entries:
        return {"error": "⚠️ Your file appears to be empty (no data rows with words found)."}

    return {
        "entries": entries,
        "mapped_headers": {v: raw_headers[k] for k, v in column_map.items()},
    }


def parse_xlsx(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    if sheet is None:
        wb.close()
        return {"error": "⚠️ Your file appears to be empty."}

    multiple_sheets = len(wb.sheetnames) > 1
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"error": "⚠️ Your file appears to be empty."}

    raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = _process_rows(raw_headers, rows[1:])
    if multiple_sheets:
        result["multiple_sheets"] = True
    return result


def parse_csv(file_path: str) -> dict:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)

    if not sample.strip():
        return {"error": "⚠️ Your file appears to be empty."}

    delimiter = _detect_csv_delimiter(sample)

    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        all_rows = list(reader)

    if not all_rows:
        return {"error": "⚠️ Your file appears to be empty."}

    raw_headers = [h.strip() for h in all_rows[0]]
    return _process_rows(raw_headers, all_rows[1:])


def parse_file(file_path: str) -> dict:
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return parse_xlsx(file_path)
    elif ext == ".csv":
        return parse_csv(file_path)
    else:
        return {"error": "❌ Unsupported file format. Please send .xlsx or .csv"}


def get_temp_path(original_filename: str) -> str:
    ext = Path(original_filename).suffix.lower()
    return os.path.join(tempfile.gettempdir(), f"ielts_vocab_{uuid.uuid4().hex}{ext}")
